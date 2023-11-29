import psycopg2 # PostreSQL 
from openpyxl import Workbook # Excel 
import os

# Get database connection parameters from environment
db_params = {
    'dbname':   os.environ.get('INGO_BENCHMARKS_DB_NAME'),
    'user':     os.environ.get('INGO_BENCHMARKS_DB_USER'),
    'password': os.environ.get('INGO_BENCHMARKS_DB_PASSWORD'),
    'host':     os.environ.get('INGO_BENCHMARKS_DB_HOST'),  
    'port':     os.environ.get('INGO_BENCHMARKS_DB_PORT') 
}

# Connect to the database
conn = psycopg2.connect(**db_params)

# Create a cursor object
cur = conn.cursor()

# Create a new workbook
wb = Workbook()
ws_mult = wb.create_sheet("Multiply") 
ws_mult.title = "Multiply"
ws_msm = wb.create_sheet("MSM") 
ws_msm.title = "MSM"
ws_msm = wb.create_sheet("Poseidon") 
ws_msm.title = "Poseidon"
ws_msm = wb.create_sheet("NTT") 
ws_msm.title = "NTT"

###############################
#     Modulo Multiply         #
###############################

# TBD

###############################
#          MSM                #
###############################

ws = wb['MSM']

# Set column headers
ws.cell(row=2, column=1, value='n')
ws.cell(row=2, column=2, value='2^n')
col_headers = 2

# Set row headers (vector size)
n_min = 10
n_max = 30

row_headers = 2

for i,n in enumerate(range(n_min, n_max)):
    ws.cell(row=row_headers+1+i, column=1, value=n)
    ws.cell(row=row_headers+1+i, column=2, value=pow(2,n))

table_name = "msm_benchmark"
device_types=['RTX 3090','RTX 4090']

for j, device_type in enumerate(device_types):
    ws_col=col_headers + 2*j + 1
    ws.cell(row=1, column=ws_col + 0, value=device_type)
    ws.cell(row=2, column=ws_col + 0, value='time (s)')
    ws.cell(row=2, column=ws_col + 1, value='power (W)')
    for i, n in enumerate(range(n_min, n_max)):
        vector_size=pow(2,n)
        query = f"""
            SELECT n.runtime_sec, n.power_Watt
            FROM {table_name} n
            JOIN hw_platform h ON n.runs_on = h.id
            WHERE h.device = '{device_type}' AND n.vector_size = '{vector_size}'
            ORDER BY n.runtime_sec ASC
            LIMIT 1;
        """

        cur.execute(query)
        result = cur.fetchone()
        if result:
            min_runtime_sec, power_watt = result
            print(f"Minimum runtime_sec is: {min_runtime_sec}")
            ws.cell(row=row_headers + i + 1, column=ws_col + 0, value=min_runtime_sec)
            print(f"Corresponding power_Watt is: {power_watt}")
            ws.cell(row=row_headers + i + 1, column=ws_col + 1, value=power_watt)
        else:
            print("No result found")


###############################
#          Poseidon           #
###############################

ws = wb['Poseidon']

# Set column headers
ws.cell(row=2, column=1, value='tree height')
# ws.cell(row=2, column=2, value='2^n')
col_headers = 2

# Set row headers (tree height)
n_min = 5
n_max = 10

row_headers = 2

for i,n in enumerate(range(n_min, n_max)):
    ws.cell(row=row_headers+1+i, column=1, value=n)
    # ws.cell(row=row_headers+1+i, column=2, value=pow(2,n))

table_name = "poseidon_benchmark"
device_types=['vu13p (U250 board)','RTX 3090','RTX 4090']

for j, device_type in enumerate(device_types):
    ws_col=col_headers + 2*j + 1
    ws.cell(row=1, column=ws_col + 0, value=device_type)
    ws.cell(row=2, column=ws_col + 0, value='time (s)')
    ws.cell(row=2, column=ws_col + 1, value='power (W)')
    for i, n in enumerate(range(n_min, n_max)):
        # vector_size=pow(2,n)
        tree_height = n
        query = f"""
            SELECT n.runtime_sec, n.power_Watt
            FROM {table_name} n
            JOIN hw_platform h ON n.runs_on = h.id
            WHERE h.device = '{device_type}' AND n.tree_height = '{tree_height}'
            ORDER BY n.runtime_sec ASC
            LIMIT 1;
        """

        cur.execute(query)
        result = cur.fetchone()
        if result:
            min_runtime_sec, power_watt = result
            print(f"Minimum runtime_sec is: {min_runtime_sec}")
            ws.cell(row=row_headers + i + 1, column=ws_col + 0, value=min_runtime_sec)
            print(f"Corresponding power_Watt is: {power_watt}")
            ws.cell(row=row_headers + i + 1, column=ws_col + 1, value=power_watt)
        else:
            print("No result found")



###############################
#          NTT                #
###############################

ws = wb['NTT']
# ws.title = "NTT"

# Set column headers
ws.cell(row=2, column=1, value='n')
ws.cell(row=2, column=2, value='2^n')
col_headers = 2

# Set row headers
n_min = 8
n_max = 28

row_headers = 2

i = 0
for n in range(n_min, n_max):
    i = i + 1
    ws.cell(row=row_headers+i, column=1, value=n)
    ws.cell(row=row_headers+i, column=2, value=pow(2,n))

table_name = "ntt_benchmark"
device_types=['RTX 3090','RTX 4090']

for j, device_type in enumerate(device_types):
    ws_col=col_headers + 2*j + 1
    ws.cell(row=1, column=ws_col + 0, value=device_type)
    ws.cell(row=2, column=ws_col + 0, value='time (s)')
    ws.cell(row=2, column=ws_col + 1, value='power (W)')
    for i, n in enumerate(range(n_min, n_max)):
        vector_size=pow(2,n)
        query = f"""
            SELECT n.runtime_sec, n.power_Watt
            FROM {table_name} n
            JOIN hw_platform h ON n.runs_on = h.id
            WHERE h.device = '{device_type}' AND n.vector_size = '{vector_size}'
            ORDER BY n.runtime_sec ASC
            LIMIT 1;
        """

        cur.execute(query)
        result = cur.fetchone()
        if result:
            min_runtime_sec, power_watt = result
            print(f"Minimum runtime_sec is: {min_runtime_sec}")
            ws.cell(row=row_headers + i + 1, column=ws_col + 0, value=min_runtime_sec)
            print(f"Corresponding power_Watt is: {power_watt}")
            ws.cell(row=row_headers + i + 1, column=ws_col + 1, value=power_watt)
        else:
            print("No result found")



# Close the cursor and connection
cur.close()
conn.close()


# Save the workbook to a file
wb.save("benchmarks.xlsx")

exit()






# You can also append rows of data
ws.append([1, 2, 3])
ws.append([4, 5, 6])

# Save the workbook to a file
wb.save("sample_workbook.xlsx")
